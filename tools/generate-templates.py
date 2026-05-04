"""
Generates Excel templates voor weekstaat, inkooporder en factuur
matching de PDF-designs (KTS huisstijl Rev A) — Rev 2 (preciezere match).

Output: templates/KTS-{Weekstaat,Inkooporder,Factuur}-template-RevA.xlsx
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# === KTS Huisstijl Rev A kleuren ===
KTS_BLUE   = "07567F"
INK_900    = "0F1B2D"
INK_700    = "2A3441"
INK_500    = "5C6675"
INK_400    = "8A93A1"
INK_300    = "B4BCC7"
LIGHT_BG   = "F8F8F4"   # cream achtergrond voor info-cellen
WEEKEND_BG = "F5F0E6"   # zandkleur weekend tint
LINE_LIGHT = "DCDCDC"

# === Logo + tandwiel paden ===
LOGO_PATH = None
for p in [
    'tools/_K_logo_temp.png',
    r'C:\Users\mkuij\OneDrive\Administratie KTS BV\Website KTDS.eu KTS\KTS Branding\KTS_Huisstijl_RevA_complete\logo_bronbestanden\png_hires\K_solo_primary_2400px.png'
]:
    if os.path.exists(p):
        LOGO_PATH = p
        break

TANDWIEL_PATH = None
for p in [
    'tools/_tandwiel_temp.png',
    'tandwiel-wit-v2.png'
]:
    if os.path.exists(p):
        TANDWIEL_PATH = p
        break

# === Helpers ===
def thin_border(color=LINE_LIGHT):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return PatternFill(start_color=KTS_BLUE, end_color=KTS_BLUE, fill_type="solid")

def soft_fill():
    return PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

def weekend_fill():
    return PatternFill(start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid")

def set_print_a4(ws, landscape=False):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if landscape else ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

def add_logo_at(ws, anchor, height_px=70):
    """Voeg KTS-logo toe op gegeven anchor cell. Aspect ratio behouden (1.5:1)."""
    if not LOGO_PATH or not os.path.exists(LOGO_PATH):
        return
    try:
        img = XLImage(LOGO_PATH)
        # K-logo is 2400x1604 (1.5:1 ratio)
        img.width = int(height_px * 1.5)
        img.height = height_px
        ws.add_image(img, anchor)
    except Exception as e:
        print(f'  logo embed faal: {e}')

def add_tandwiel_at(ws, anchor, height_px=24):
    """Voeg klein tandwiel-icoontje toe naast de titel (zoals in PDF)."""
    if not TANDWIEL_PATH or not os.path.exists(TANDWIEL_PATH):
        return
    try:
        img = XLImage(TANDWIEL_PATH)
        img.width = height_px
        img.height = height_px
        ws.add_image(img, anchor)
    except Exception as e:
        print(f'  tandwiel embed faal: {e}')

def set_cell(ws, ref, value, font_kwargs=None, fill=None, align=None, border=None, num_format=None):
    """Helper: set cell met optionele styling"""
    c = ws[ref] if isinstance(ref, str) else ws.cell(row=ref[0], column=ref[1])
    c.value = value
    if font_kwargs:
        c.font = Font(**font_kwargs)
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c

# =============================================================
# 1. WEEKSTAAT — landscape A4
#    Layout 1-op-1 met generateWeekstaatV2() PDF-design
# =============================================================
def make_weekstaat():
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekstaat"
    set_print_a4(ws, landscape=True)

    # Kolombreedtes (units = ~7px elk)
    # PDF: DAG 22mm/DATUM 22/BEGIN 18/EIND 18/PAUZE 16/UREN 20/WERK 98/LOC 42/KM 17 = 273mm
    # Ratio behouden: 273 → ~140 column units totaal voor A4 landscape
    widths = [11, 11, 9, 9, 8, 10, 50, 22, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ====== HEADER ROW: tandwiel + WEEKSTAAT (links) | logo + adres (rechts) ======
    # Tandwiel-icoon op A1 (klein)
    add_tandwiel_at(ws, 'A1', height_px=22)
    # Titel WEEKSTAAT in A1, met whitespace voor icoon
    ws['A1'] = "  WEEKSTAAT"
    ws['A1'].font = Font(name='Calibri', size=22, bold=True, color=KTS_BLUE)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # Logo rechtsboven (kolommen H-I)
    add_logo_at(ws, 'H1', height_px=60)

    # Adres rechts onder logo (rij 5-6)
    set_cell(ws, 'F5', "Nieuwboerweg 2A, 1738BB Waarland",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('F5:I5')
    set_cell(ws, 'F6', "+31 6 5123 9050  ·  info@kuijpers-ts.nl",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('F6:I6')

    # ====== INFO BAR LINKS: NAAM | PERIODE  (rij 3-4)  ======
    set_cell(ws, 'A3', "NAAM",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400})
    set_cell(ws, 'D3', "PERIODE",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400})

    set_cell(ws, 'A4', "[Vul je naam in]",
        font_kwargs={'name':'Calibri','size':12,'bold':True,'color':INK_900})
    ws.merge_cells('A4:C4')
    set_cell(ws, 'D4', "Week XX · 2026",
        font_kwargs={'name':'Calibri','size':12,'bold':True,'color':INK_900})
    ws.merge_cells('D4:E4')

    # PROJECT | OPDRACHTGEVER  (rij 6-7)
    set_cell(ws, 'A6', "PROJECT",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400})
    set_cell(ws, 'D6', "OPDRACHTGEVER",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400})

    set_cell(ws, 'A7', "[Projectnaam]",
        font_kwargs={'name':'Calibri','size':11,'bold':True,'color':INK_900})
    ws.merge_cells('A7:C7')
    set_cell(ws, 'D7', "[Klantnaam]",
        font_kwargs={'name':'Calibri','size':11,'bold':True,'color':INK_900})
    ws.merge_cells('D7:E7')

    # ====== TABEL HEADER (rij 9) ======
    headers = [
        ("DAG",          'left'),
        ("DATUM",        'center'),
        ("BEGINTIJD",    'center'),
        ("EINDTIJD",     'center'),
        ("PAUZE",        'center'),
        ("GEW. UREN",    'center'),
        ("WERKZAAMHEDEN",'left'),
        ("LOCATIE",      'left'),
        ("KM",           'center')
    ]
    HEADER_ROW = 9
    for col, (h, align) in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
        c.fill = header_fill()
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border = thin_border()
    ws.row_dimensions[HEADER_ROW].height = 22

    # ====== DAG-RIJEN (rij 10-16, 7 dagen) ======
    days = ['Maandag','Dinsdag','Woensdag','Donderdag','Vrijdag','Zaterdag','Zondag']
    for i, day in enumerate(days):
        row = HEADER_ROW + 1 + i
        is_weekend = i >= 5
        # DAG cel (links uitgelijnd)
        c = ws.cell(row=row, column=1, value=day)
        c.font = Font(name='Calibri', size=10, color=INK_900)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        # Borders + alignment per kolom
        for col, (_, align) in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal=align, vertical='center',
                                       indent=1 if align == 'left' else 0)
            if is_weekend:
                cell.fill = weekend_fill()
            cell.font = cell.font.copy() if cell.value else Font(name='Calibri', size=10, color=INK_900)
        ws.row_dimensions[row].height = 22

    # ====== KPI-STRIP (rij 18-19) ======
    KPI_LABEL_ROW = HEADER_ROW + 9   # rij 18
    KPI_VAL_ROW   = KPI_LABEL_ROW + 1

    kpi_data = [
        ("REGULIER MA-VR", f"=SUM(F{HEADER_ROW+1}:F{HEADER_ROW+5})", False),
        ("ZATERDAG",       f"=F{HEADER_ROW+6}",                     False),
        ("ZONDAG/FEEST",   f"=F{HEADER_ROW+7}",                     False),
        ("REIS KM",        f"=SUM(I{HEADER_ROW+1}:I{HEADER_ROW+7})",False),
        ("TOTAAL TE FACTUREREN", f"=SUM(F{HEADER_ROW+1}:F{HEADER_ROW+7})", True),
    ]
    # KPIs occuperen kolommen 1-5 elk 1 kolom (samen 5 kolommen)
    for i, (label, formula, accent) in enumerate(kpi_data):
        col = i + 1
        # Label cel
        lc = ws.cell(row=KPI_LABEL_ROW, column=col, value=label)
        lc.font = Font(name='Calibri', size=8, bold=accent,
                       color="FFFFFF" if accent else INK_400)
        lc.fill = header_fill() if accent else soft_fill()
        lc.alignment = Alignment(horizontal='center', vertical='center')
        lc.border = thin_border()
        # Value cel
        vc = ws.cell(row=KPI_VAL_ROW, column=col, value=formula)
        vc.font = Font(name='Calibri', size=20, bold=True,
                       color="FFFFFF" if accent else INK_900)
        vc.fill = header_fill() if accent else soft_fill()
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.border = thin_border()
    ws.row_dimensions[KPI_LABEL_ROW].height = 16
    ws.row_dimensions[KPI_VAL_ROW].height = 32

    # ====== OPMERKINGEN (rechts naast KPI-strip, kolommen 6-9) ======
    set_cell(ws, (KPI_LABEL_ROW, 6), "OPMERKINGEN",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400},
        fill=soft_fill(),
        align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.merge_cells(start_row=KPI_LABEL_ROW, start_column=6,
                   end_row=KPI_LABEL_ROW, end_column=9)
    # Borders + soft fill voor opmerkingen-area
    for col in range(6, 10):
        ws.cell(row=KPI_LABEL_ROW, column=col).border = thin_border()
        ws.cell(row=KPI_LABEL_ROW, column=col).fill = soft_fill()
        ws.cell(row=KPI_VAL_ROW, column=col).border = thin_border()
        ws.cell(row=KPI_VAL_ROW, column=col).fill = soft_fill()
    ws.merge_cells(start_row=KPI_VAL_ROW, start_column=6,
                   end_row=KPI_VAL_ROW, end_column=9)
    ws.cell(row=KPI_VAL_ROW, column=6).alignment = Alignment(
        horizontal='left', vertical='top', wrap_text=True, indent=1)

    # ====== HANDTEKENING-BOXEN (2 onder elkaar of naast elkaar) ======
    SIG_LABEL_ROW = KPI_VAL_ROW + 2
    set_cell(ws, (SIG_LABEL_ROW, 1), "HANDTEKENING OPDRACHTNEMER",
        font_kwargs={'name':'Calibri','size':8,'bold':True,'color':INK_400},
        align=Alignment(horizontal='left', indent=1))
    ws.merge_cells(start_row=SIG_LABEL_ROW, start_column=1,
                   end_row=SIG_LABEL_ROW, end_column=4)
    set_cell(ws, (SIG_LABEL_ROW, 6), "HANDTEKENING OPDRACHTGEVER",
        font_kwargs={'name':'Calibri','size':8,'bold':True,'color':INK_400},
        align=Alignment(horizontal='left', indent=1))
    ws.merge_cells(start_row=SIG_LABEL_ROW, start_column=6,
                   end_row=SIG_LABEL_ROW, end_column=9)

    # Box-rijen (3 rows hoog)
    for r in range(SIG_LABEL_ROW, SIG_LABEL_ROW + 4):
        for col_range in [(1, 4), (6, 9)]:
            for col in range(col_range[0], col_range[1] + 1):
                ws.cell(row=r, column=col).border = thin_border()
        if r > SIG_LABEL_ROW:
            ws.row_dimensions[r].height = 24

    # ====== FOOTER ======
    FOOT_ROW = SIG_LABEL_ROW + 5
    set_cell(ws, (FOOT_ROW, 1),
        "Op deze opdracht zijn de Algemene Voorwaarden Detachering 2026 van Kuijpers Technical Services BV van toepassing.",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400})
    ws.merge_cells(start_row=FOOT_ROW, start_column=1,
                   end_row=FOOT_ROW, end_column=7)
    set_cell(ws, (FOOT_ROW, 8), "KvK 93410557  ·  BTW NL866385368B01",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400},
        align=Alignment(horizontal='right'))
    ws.merge_cells(start_row=FOOT_ROW, start_column=8,
                   end_row=FOOT_ROW, end_column=9)

    out = 'templates/KTS-Weekstaat-template-RevA.xlsx'
    wb.save(out)
    print(f'  Weekstaat: {out}')


# =============================================================
# 2. INKOOPORDER — portrait A4
# =============================================================
def make_inkooporder():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inkooporder"
    set_print_a4(ws, landscape=False)

    # PDF: 7 cols voor info-bar + items-tabel
    # Items: ITEM# 18 / OMSCHR 92 / AANT 18 / PRIJS 28 / TOTAAL 30 = 186mm
    widths = [8, 30, 18, 14, 16, 22, 8]   # 7 kolommen
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ====== HEADER: tandwiel + INKOOPORDER (links), logo + adres (rechts) ======
    add_tandwiel_at(ws, 'A1', height_px=22)
    ws['A1'] = "  INKOOPORDER"
    ws['A1'].font = Font(name='Calibri', size=22, bold=True, color=KTS_BLUE)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:D1')

    # Logo rechtsboven (kolommen E-G)
    add_logo_at(ws, 'F1', height_px=56)

    # Adres rechts onder logo
    set_cell(ws, 'E5', "Nieuwboerweg 2A, 1738BB Waarland",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('E5:G5')
    set_cell(ws, 'E6', "+31 6 5123 9050  ·  info@kuijpers-ts.nl",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('E6:G6')

    # ====== INFO BAR: 4 cellen (rij 8-9) ======
    INFO_ROW = 8
    info_cells = [
        ("PROJECT",         "[Projectcode]", 1, 2),
        ("SCHEIDINGSTEKEN", "-",            3, 3),
        ("DATUM",           "[DD-MM-JJJJ]", 4, 5),
        ("PO-NUMMER",       "[Auto]",       6, 7),
    ]
    for label, value, sc, ec in info_cells:
        set_cell(ws, (INFO_ROW, sc), label,
            font_kwargs={'name':'Calibri','size':8,'color':INK_400},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        set_cell(ws, (INFO_ROW + 1, sc), value,
            font_kwargs={'name':'Calibri','size':11,'bold':True,'color':INK_900},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        if sc != ec:
            ws.merge_cells(start_row=INFO_ROW, start_column=sc, end_row=INFO_ROW, end_column=ec)
            ws.merge_cells(start_row=INFO_ROW + 1, start_column=sc, end_row=INFO_ROW + 1, end_column=ec)
        for r in [INFO_ROW, INFO_ROW + 1]:
            for col in range(sc, ec + 1):
                ws.cell(row=r, column=col).fill = soft_fill()
                ws.cell(row=r, column=col).border = thin_border()
    ws.row_dimensions[INFO_ROW].height = 14
    ws.row_dimensions[INFO_ROW + 1].height = 22

    # ====== LEVERANCIER + LEVERADRES blokken (rij 11+) ======
    BLOCK_HEADER = INFO_ROW + 3
    # Leverancier header (links, kolommen 1-3)
    set_cell(ws, (BLOCK_HEADER, 1), "LEVERANCIER",
        font_kwargs={'name':'Calibri','size':9,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.merge_cells(start_row=BLOCK_HEADER, start_column=1,
                   end_row=BLOCK_HEADER, end_column=3)
    for col in range(1, 4):
        ws.cell(row=BLOCK_HEADER, column=col).fill = header_fill()
        ws.cell(row=BLOCK_HEADER, column=col).border = thin_border()
    # Leveradres header (rechts, kolommen 4-7)
    set_cell(ws, (BLOCK_HEADER, 4), "LEVERADRES",
        font_kwargs={'name':'Calibri','size':9,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.merge_cells(start_row=BLOCK_HEADER, start_column=4,
                   end_row=BLOCK_HEADER, end_column=7)
    for col in range(4, 8):
        ws.cell(row=BLOCK_HEADER, column=col).fill = header_fill()
        ws.cell(row=BLOCK_HEADER, column=col).border = thin_border()
    ws.row_dimensions[BLOCK_HEADER].height = 18

    # Block content
    leverancier_data = [
        ("[Bedrijfsnaam leverancier]", True),
        ("[Contactpersoon]", False),
        ("[Adres]", False),
        ("[Postcode + plaats]", False),
        ("[Telefoon]", False),
        ("[E-mail]", False),
    ]
    leveradres_data = [
        ("Kuijpers Technical Services BV", True),
        ("Crediteurenadministratie", False),
        ("Nieuwboerweg 2A", False),
        ("1738BB, Waarland", False),
        ("+31 6 5123 9050", False),
        ("info@kuijpers-ts.nl", False),
    ]
    for i, ((lev_text, lev_bold), (lvr_text, lvr_bold)) in enumerate(zip(leverancier_data, leveradres_data)):
        r = BLOCK_HEADER + 1 + i
        # Leverancier (kolommen 1-3)
        set_cell(ws, (r, 1), lev_text,
            font_kwargs={'name':'Calibri','size':9,'bold':lev_bold,
                         'color':INK_900 if lev_bold else INK_500},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        # Leveradres (kolommen 4-7)
        set_cell(ws, (r, 4), lvr_text,
            font_kwargs={'name':'Calibri','size':9,'bold':lvr_bold,
                         'color':INK_900 if lvr_bold else INK_500},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        # Borders + soft fill
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = soft_fill()
            ws.cell(row=r, column=col).border = thin_border()
        ws.row_dimensions[r].height = 16

    # ====== ITEMS TABEL ======
    ITEMS_HEADER_ROW = BLOCK_HEADER + 8
    item_headers = [
        ("ITEM #",       1, 1, 'center'),
        ("OMSCHRIJVING", 2, 3, 'left'),
        ("AANTAL",       4, 4, 'center'),
        ("PRIJS / STUK", 5, 5, 'right'),
        ("TOTAAL",       6, 7, 'right'),
    ]
    for label, sc, ec, align in item_headers:
        c = ws.cell(row=ITEMS_HEADER_ROW, column=sc, value=label)
        c.font = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
        c.fill = header_fill()
        c.alignment = Alignment(horizontal=align, vertical='center',
                                indent=1 if align == 'left' else 0)
        if sc != ec:
            ws.merge_cells(start_row=ITEMS_HEADER_ROW, start_column=sc,
                           end_row=ITEMS_HEADER_ROW, end_column=ec)
        for col in range(sc, ec + 1):
            ws.cell(row=ITEMS_HEADER_ROW, column=col).fill = header_fill()
            ws.cell(row=ITEMS_HEADER_ROW, column=col).border = thin_border()
    ws.row_dimensions[ITEMS_HEADER_ROW].height = 22

    # 8 lege rijen (2-regel desc support via wrap_text + extra hoogte)
    for i in range(1, 9):
        r = ITEMS_HEADER_ROW + i
        # ITEM #
        ws.cell(row=r, column=1, value=i).font = Font(name='Calibri', size=10, color=INK_500)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
        # OMSCHRIJVING (kolommen 2-3, merged, wrap_text)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True, indent=1)
        ws.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=INK_900)
        # AANTAL
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='center')
        # PRIJS
        ws.cell(row=r, column=5).number_format = '€ #,##0.00'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='right', vertical='center', indent=1)
        # TOTAAL (kolommen 6-7, merged, formule)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.cell(row=r, column=6, value=f"=IFERROR(D{r}*E{r},\"\")").number_format = '€ #,##0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws.cell(row=r, column=6).font = Font(name='Calibri', size=10, bold=True, color=INK_900)
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border()
        ws.row_dimensions[r].height = 24  # ruimte voor 2-regel desc

    # ====== OPMERKINGEN-BLOK + TOTALEN-BLOK ======
    BOT_HEADER = ITEMS_HEADER_ROW + 9 + 1   # + 8 items + 1 spacer

    # Opmerkingen header (links, kolommen 1-4)
    set_cell(ws, (BOT_HEADER, 1), "OPMERKINGEN OF SPECIALE INSTRUCTIES",
        font_kwargs={'name':'Calibri','size':8,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='left', vertical='center', indent=1))
    ws.merge_cells(start_row=BOT_HEADER, start_column=1, end_row=BOT_HEADER, end_column=4)
    for col in range(1, 5):
        ws.cell(row=BOT_HEADER, column=col).fill = header_fill()
        ws.cell(row=BOT_HEADER, column=col).border = thin_border()
    ws.row_dimensions[BOT_HEADER].height = 18

    # Opmerkingen body (3 rijen)
    for i in range(1, 4):
        r = BOT_HEADER + i
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = soft_fill()
            ws.cell(row=r, column=col).border = thin_border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        if i == 1:
            ws.cell(row=r, column=1, value="[Vul opmerkingen in...]")
            ws.cell(row=r, column=1).font = Font(name='Calibri', size=10, color=INK_500)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        ws.row_dimensions[r].height = 16

    # Betalingstermijn onderaan opmerkingen
    BETALING_ROW = BOT_HEADER + 4
    for col in range(1, 5):
        ws.cell(row=BETALING_ROW, column=col).fill = soft_fill()
        ws.cell(row=BETALING_ROW, column=col).border = thin_border()
    ws.merge_cells(start_row=BETALING_ROW, start_column=1, end_row=BETALING_ROW, end_column=4)
    set_cell(ws, (BETALING_ROW, 1), "BETALINGSTERMIJN",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400},
        align=Alignment(horizontal='left', vertical='center', indent=1))
    BETALING_VAL_ROW = BETALING_ROW + 1
    for col in range(1, 5):
        ws.cell(row=BETALING_VAL_ROW, column=col).fill = soft_fill()
        ws.cell(row=BETALING_VAL_ROW, column=col).border = thin_border()
    ws.merge_cells(start_row=BETALING_VAL_ROW, start_column=1, end_row=BETALING_VAL_ROW, end_column=4)
    set_cell(ws, (BETALING_VAL_ROW, 1), "30 dagen na factuurdatum",
        font_kwargs={'name':'Calibri','size':10,'bold':True,'color':INK_900},
        align=Alignment(horizontal='left', vertical='center', indent=1))

    # Totalen-blok (rechts, kolommen 5-7)
    sub_formula = f"=SUM(F{ITEMS_HEADER_ROW+1}:F{ITEMS_HEADER_ROW+8})"
    tot_rows = [
        ("Subtotaal",          sub_formula),
        ("BTW (21%)",          f"=F{BOT_HEADER}*0.21"),  # placeholder; gefixed na adjustment
        ("Transport",          0),
        ("Overige",            0),
    ]
    for i, (label, formula) in enumerate(tot_rows):
        r = BOT_HEADER + i
        set_cell(ws, (r, 5), label,
            font_kwargs={'name':'Calibri','size':10,'color':INK_500},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        # Set formule met juiste row-referentie
        if i == 0:
            ws.cell(row=r, column=6, value=sub_formula)
        elif i == 1:
            ws.cell(row=r, column=6, value=f"=F{BOT_HEADER}*0.21")
        else:
            ws.cell(row=r, column=6, value=formula)
        ws.cell(row=r, column=6).number_format = '€ #,##0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws.cell(row=r, column=6).font = Font(name='Calibri', size=10, bold=True, color=INK_900)
        ws.cell(row=r, column=6).fill = soft_fill()
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        for col in range(5, 8):
            ws.cell(row=r, column=col).fill = soft_fill()
            ws.cell(row=r, column=col).border = thin_border()
        ws.row_dimensions[r].height = 18

    # TOTAAL-band (rechts onderaan)
    TOT_BAND_ROW = BOT_HEADER + 4
    set_cell(ws, (TOT_BAND_ROW, 5), "TOTAAL",
        font_kwargs={'name':'Calibri','size':12,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='left', vertical='center', indent=1))
    set_cell(ws, (TOT_BAND_ROW, 6),
        f"=F{BOT_HEADER}+F{BOT_HEADER+1}+F{BOT_HEADER+2}+F{BOT_HEADER+3}",
        font_kwargs={'name':'Calibri','size':14,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='right', vertical='center', indent=1),
        num_format='€ #,##0.00')
    ws.merge_cells(start_row=TOT_BAND_ROW, start_column=6,
                   end_row=TOT_BAND_ROW, end_column=7)
    for col in range(5, 8):
        ws.cell(row=TOT_BAND_ROW, column=col).fill = header_fill()
        ws.cell(row=TOT_BAND_ROW, column=col).border = thin_border()
    ws.row_dimensions[TOT_BAND_ROW].height = 26

    # ====== FOOTER ======
    FOOT_ROW = TOT_BAND_ROW + 3
    set_cell(ws, (FOOT_ROW, 1),
        "Op deze inkooporder zijn de Algemene Voorwaarden 2026 van Kuijpers Technical Services BV van toepassing.",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400})
    ws.merge_cells(start_row=FOOT_ROW, start_column=1, end_row=FOOT_ROW, end_column=4)
    set_cell(ws, (FOOT_ROW, 5), "KvK 93410557  ·  BTW NL866385368B01",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400},
        align=Alignment(horizontal='right'))
    ws.merge_cells(start_row=FOOT_ROW, start_column=5, end_row=FOOT_ROW, end_column=7)

    out = 'templates/KTS-Inkooporder-template-RevA.xlsx'
    wb.save(out)
    print(f'  Inkooporder: {out}')


# =============================================================
# 3. FACTUUR — portrait A4
# =============================================================
def make_factuur():
    wb = Workbook()
    ws = wb.active
    ws.title = "Factuur"
    set_print_a4(ws, landscape=False)

    # PDF: PERIODE 28 / OMSCHR 78 / AANT 22 / TARIEF 22 / SUBTOT 24 / BTW 12 = 186mm
    widths = [13, 32, 11, 11, 13, 8, 4]   # 7 kolommen incl spacer
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ====== HEADER: tandwiel + FACTUUR (links) | logo + adres (rechts) ======
    add_tandwiel_at(ws, 'A1', height_px=22)
    ws['A1'] = "  FACTUUR"
    ws['A1'].font = Font(name='Calibri', size=22, bold=True, color=KTS_BLUE)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:D1')

    add_logo_at(ws, 'F1', height_px=56)

    set_cell(ws, 'E5', "Nieuwboerweg 2A, 1738BB Waarland",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('E5:G5')
    set_cell(ws, 'E6', "+31 6 5123 9050  ·  info@kuijpers-ts.nl",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500},
        align=Alignment(horizontal='right'))
    ws.merge_cells('E6:G6')

    # ====== KLANT-BLOK (onder titel, links) ======
    set_cell(ws, 'A4', "AAN",
        font_kwargs={'name':'Calibri','size':8,'color':INK_400})
    set_cell(ws, 'A5', "[Klantnaam B.V.]",
        font_kwargs={'name':'Calibri','size':12,'bold':True,'color':INK_900})
    ws.merge_cells('A5:C5')
    klant_lines = [
        "[t.a.v. crediteurenadministratie]",
        "[invoices@klant.nl]",
        "[Postbus 800]",
        "[2800 AB Gouda]",
    ]
    for i, line in enumerate(klant_lines):
        r = 6 + i
        set_cell(ws, (r, 1), line,
            font_kwargs={'name':'Calibri','size':10,'color':INK_500})
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    # ====== INFO BAR: 4 cellen (rij 11-12) ======
    INFO_ROW = 11
    info_cells = [
        ("FACTUURNUMMER", "[2026-XX]",     1, 2),
        ("FACTUURDATUM",  "[DD-MM-JJJJ]",  3, 3),
        ("VERVALDATUM",   "[DD-MM-JJJJ]",  4, 4),
        ("PROJECT",       "[Projectnaam]", 5, 7),
    ]
    for label, value, sc, ec in info_cells:
        set_cell(ws, (INFO_ROW, sc), label,
            font_kwargs={'name':'Calibri','size':8,'color':INK_400},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1))
        font_size = 10 if label == 'PROJECT' else 11
        set_cell(ws, (INFO_ROW + 1, sc), value,
            font_kwargs={'name':'Calibri','size':font_size,'bold':True,'color':INK_900},
            fill=soft_fill(),
            align=Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True))
        if sc != ec:
            ws.merge_cells(start_row=INFO_ROW, start_column=sc, end_row=INFO_ROW, end_column=ec)
            ws.merge_cells(start_row=INFO_ROW + 1, start_column=sc, end_row=INFO_ROW + 1, end_column=ec)
        for r in [INFO_ROW, INFO_ROW + 1]:
            for col in range(sc, ec + 1):
                ws.cell(row=r, column=col).fill = soft_fill()
                ws.cell(row=r, column=col).border = thin_border()
    ws.row_dimensions[INFO_ROW].height = 14
    ws.row_dimensions[INFO_ROW + 1].height = 22

    # ====== PROJECT-INFO compact + Loonheffingen ======
    PROJ_ROW = INFO_ROW + 3
    proj_info = [
        ("Projectnummer",   "[Nummer]", False),
        ("Opdrachtnummer",  "[Nummer]", False),
        ("PO-nummer",       "[Nummer]", False),
    ]
    for i, (lbl, val, _) in enumerate(proj_info):
        r = PROJ_ROW + i
        set_cell(ws, (r, 1), lbl,
            font_kwargs={'name':'Calibri','size':9,'color':INK_500})
        set_cell(ws, (r, 3), val,
            font_kwargs={'name':'Calibri','size':9,'color':INK_900})
        ws.row_dimensions[r].height = 14

    # Loonheffingennummers (klein, niet-bold)
    LOON_ROW = PROJ_ROW + 3
    loon_lines = [
        ("Loonheffingennummer KTDS Holding B.V.",         "866381557L01"),
        ("Loonheffingennummer Kuijpers TD Holding B.V.",  "866381594L01"),
    ]
    for i, (lbl, val) in enumerate(loon_lines):
        r = LOON_ROW + i
        set_cell(ws, (r, 1), lbl,
            font_kwargs={'name':'Calibri','size':8,'color':INK_400})
        set_cell(ws, (r, 4), val,
            font_kwargs={'name':'Calibri','size':8,'color':INK_500})
        ws.row_dimensions[r].height = 12

    # ====== ITEMS TABEL ======
    ITEMS_HEADER_ROW = LOON_ROW + 3
    item_headers = [
        ("PERIODE",      'left'),
        ("OMSCHRIJVING", 'left'),
        ("AANTAL",       'right'),
        ("TARIEF",       'right'),
        ("SUBTOTAAL",    'right'),
        ("BTW %",        'right'),
    ]
    for col, (h, align) in enumerate(item_headers, start=1):
        c = ws.cell(row=ITEMS_HEADER_ROW, column=col, value=h)
        c.font = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
        c.fill = header_fill()
        c.alignment = Alignment(horizontal=align, vertical='center',
                                indent=1 if align == 'left' else 0)
        c.border = thin_border()
    ws.row_dimensions[ITEMS_HEADER_ROW].height = 22

    # 6 lege rijen
    for i in range(1, 7):
        r = ITEMS_HEADER_ROW + i
        # SUBTOTAAL formule
        ws.cell(row=r, column=5, value=f"=IFERROR(C{r}*D{r},\"\")")
        ws.cell(row=r, column=5).number_format = '€ #,##0.00'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws.cell(row=r, column=5).font = Font(name='Calibri', size=10, bold=True, color=INK_900)
        # TARIEF format
        ws.cell(row=r, column=4).number_format = '€ #,##0.00'
        # BTW% default 21
        ws.cell(row=r, column=6, value=21).alignment = Alignment(horizontal='right', vertical='center', indent=1)
        # Borders + alignment per kolom
        for col, (_, align) in enumerate(item_headers, start=1):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border()
            if not cell.alignment.horizontal:
                cell.alignment = Alignment(horizontal=align, vertical='center',
                                           indent=1 if align == 'left' else 0,
                                           wrap_text=(col == 2))
            cell.font = cell.font.copy() if cell.value else Font(name='Calibri', size=10, color=INK_900)
        ws.row_dimensions[r].height = 22

    # ====== TOTALEN-BLOK rechts onderaan ======
    TOT_ROW = ITEMS_HEADER_ROW + 8
    sub_formula = f"=SUM(E{ITEMS_HEADER_ROW+1}:E{ITEMS_HEADER_ROW+6})"

    set_cell(ws, (TOT_ROW, 4), "Subtotaal excl. BTW",
        font_kwargs={'name':'Calibri','size':10,'color':INK_500},
        align=Alignment(horizontal='right', indent=1))
    set_cell(ws, (TOT_ROW, 5), sub_formula,
        font_kwargs={'name':'Calibri','size':10,'bold':True,'color':INK_900},
        align=Alignment(horizontal='right', indent=1),
        num_format='€ #,##0.00')

    set_cell(ws, (TOT_ROW + 1, 4), "BTW 21%",
        font_kwargs={'name':'Calibri','size':10,'color':INK_500},
        align=Alignment(horizontal='right', indent=1))
    set_cell(ws, (TOT_ROW + 1, 5), f"=E{TOT_ROW}*0.21",
        font_kwargs={'name':'Calibri','size':10,'bold':True,'color':INK_900},
        align=Alignment(horizontal='right', indent=1),
        num_format='€ #,##0.00')

    # TOTAAL TE BETALEN band
    TOT_BAND = TOT_ROW + 2
    set_cell(ws, (TOT_BAND, 4), "TOTAAL TE BETALEN",
        font_kwargs={'name':'Calibri','size':11,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='right', vertical='center', indent=1))
    set_cell(ws, (TOT_BAND, 5), f"=E{TOT_ROW}+E{TOT_ROW+1}",
        font_kwargs={'name':'Calibri','size':14,'bold':True,'color':"FFFFFF"},
        fill=header_fill(),
        align=Alignment(horizontal='right', vertical='center', indent=1),
        num_format='€ #,##0.00')
    for col in range(4, 7):
        ws.cell(row=TOT_BAND, column=col).fill = header_fill()
    ws.row_dimensions[TOT_BAND].height = 26

    # ====== BETAAL-VERZOEK + IBAN ======
    PAY_ROW = TOT_BAND + 4
    set_cell(ws, (PAY_ROW, 1),
        "Wij verzoeken u vriendelijk het totaalbedrag uiterlijk vervaldatum over te maken,",
        font_kwargs={'name':'Calibri','size':10,'color':INK_900})
    ws.merge_cells(start_row=PAY_ROW, start_column=1, end_row=PAY_ROW, end_column=7)
    set_cell(ws, (PAY_ROW + 1, 1),
        "onder vermelding van het factuurnummer, naar onderstaande bankrekening:",
        font_kwargs={'name':'Calibri','size':10,'color':INK_900})
    ws.merge_cells(start_row=PAY_ROW + 1, start_column=1, end_row=PAY_ROW + 1, end_column=7)
    set_cell(ws, (PAY_ROW + 3, 1), "IBAN  NL61 BUNQ 2113 3747 30",
        font_kwargs={'name':'Calibri','size':12,'bold':True,'color':KTS_BLUE})
    ws.merge_cells(start_row=PAY_ROW + 3, start_column=1, end_row=PAY_ROW + 3, end_column=7)
    set_cell(ws, (PAY_ROW + 4, 1), "t.n.v. Kuijpers Technical Services BV  ·  BIC: BUNQNL2A",
        font_kwargs={'name':'Calibri','size':9,'color':INK_500})
    ws.merge_cells(start_row=PAY_ROW + 4, start_column=1, end_row=PAY_ROW + 4, end_column=7)

    # ====== FOOTER ======
    FOOT_ROW = PAY_ROW + 7
    set_cell(ws, (FOOT_ROW, 1),
        "Op deze factuur zijn de Algemene Voorwaarden 2026 van Kuijpers Technical Services BV van toepassing.",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400})
    ws.merge_cells(start_row=FOOT_ROW, start_column=1, end_row=FOOT_ROW, end_column=4)
    set_cell(ws, (FOOT_ROW, 5), "KvK 93410557  ·  BTW NL866385368B01",
        font_kwargs={'name':'Calibri','size':7,'color':INK_400},
        align=Alignment(horizontal='right'))
    ws.merge_cells(start_row=FOOT_ROW, start_column=5, end_row=FOOT_ROW, end_column=7)

    out = 'templates/KTS-Factuur-template-RevA.xlsx'
    wb.save(out)
    print(f'  Factuur: {out}')


if __name__ == '__main__':
    print('Genereren KTS Excel templates Rev A (preciezere PDF-match):')
    make_weekstaat()
    make_inkooporder()
    make_factuur()
    print('\nKlaar! 3 templates in templates/ folder.')
