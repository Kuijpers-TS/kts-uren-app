"""
Converter: KTS Onderhoudsconcept Excel -> KTS Inspectie App import-Excel

Leest het 'Onderhoudsconcept' sheet uit een leverancier-Excel (zoals
ASD-REG-0400-0.01) en genereert een Excel die direct geimporteerd
kan worden via Beheer -> Formulieren -> Importeer Excel.

Filtert op inspectie-taken (Taakomschrijving start met 'Inspecteer'
of 'Controleer'). Maakt 1 template per generiek element, met sectie
per specifiek element / bouwdeel.

Gebruik:
    python docs/convert_ohc_to_kts.py <input.xlsx> [<output.xlsx>]

Standaard output: docs/KTS_Inspectie_uit_OHC.xlsx
"""

import sys
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def detect_question_type(taakomschrijving: str) -> str:
    """Heuristiek: welk vraagtype past het best."""
    t = (taakomschrijving or '').lower()
    if any(w in t for w in ['meet ', 'meet de ', 'controleer de waarde', 'analyse', 'waarde']):
        return 'meting'
    if t.startswith('documenteer') or 'noteer' in t:
        return 'tekst'
    if 'controleer' in t and ('werking' in t or 'aanslui' in t or 'losse' in t or 'juiste' in t):
        return 'goed_fout'
    # Default voor visuele inspecties: conditiescore (NEN 2767)
    return 'conditiescore'


def detect_unit(taakomschrijving: str) -> str:
    """Probeer een eenheid af te leiden voor metingen."""
    t = (taakomschrijving or '').lower()
    if 'temperatuur' in t: return '°C'
    if 'druk' in t and 'val' in t: return 'bar'
    if 'spanning' in t: return 'V'
    if 'stroom' in t: return 'A'
    if 'isolatieweerstand' in t or 'megger' in t: return 'MΩ'
    if 'trilling' in t: return 'mm/s'
    if 'zuurgraad' in t or 'ph' in t: return 'pH'
    if 'frequentie' in t: return 'Hz'
    return ''


def section_title_from_row(row: dict) -> str:
    """Bouw sectie-titel uit element + bouwdeel."""
    elem = (row.get('Element Specifiek') or row.get('Element Generiek') or '').strip()
    bouwdeel = (row.get('Bouwdeel specifiek') or row.get('Bouwdeel generiek') or '').strip()
    if bouwdeel and bouwdeel.lower() not in elem.lower():
        return f"{elem} - {bouwdeel}"
    return elem or 'Algemeen'


def normalize(s: str) -> str:
    return re.sub(r'\s+', ' ', (s or '').strip())


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'docs/KTS_Inspectie_uit_OHC.xlsx'

    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    if 'Onderhoudsconcept' not in wb_in.sheetnames:
        print(f"FOUT: sheet 'Onderhoudsconcept' niet gevonden in {input_path}")
        sys.exit(1)

    ws = wb_in['Onderhoudsconcept']
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    print(f"Ingelezen: {len(rows)} rijen")

    # Filter inspectie-taken
    insp_rows = [r for r in rows if (r.get('Taakomschrijving') or '').strip().lower().startswith(('inspecteer', 'controleer'))]
    print(f"Inspectie-taken: {len(insp_rows)}")

    # Groepeer per Element Generiek -> sectie -> taken
    templates = defaultdict(lambda: defaultdict(list))
    for r in insp_rows:
        elem = (r.get('Element Generiek') or '').strip()
        if not elem:
            continue
        sec_title = section_title_from_row(r)
        templates[elem][sec_title].append(r)

    print(f"\nTemplates: {len(templates)}")
    for tpl, secs in templates.items():
        total = sum(len(qs) for qs in secs.values())
        print(f"  {tpl}: {len(secs)} secties, {total} vragen")

    # === Bouw output Excel ===
    wb_out = openpyxl.Workbook()

    # Sheet: Instructies
    ws_i = wb_out.active
    ws_i.title = 'Instructies'
    ws_i['A1'] = 'KTS Inspectie templates uit onderhoudsconcept'
    ws_i['A1'].font = Font(name='Arial', size=14, bold=True, color='07567F')
    instr = [
        '',
        f'Bron: {Path(input_path).name}',
        f'Aantal templates: {len(templates)}',
        f'Aantal vragen: {sum(sum(len(qs) for qs in secs.values()) for secs in templates.values())}',
        '',
        'WERKWIJZE:',
        '1. Open sheet "Vragen" en check de gegenereerde vragen',
        '2. Pas vraagtypes aan waar nodig (auto-detectie kan ernaast zitten)',
        '3. Verwijder of voeg vragen toe',
        '4. Sla op',
        '5. Importeer in app: Beheer -> Formulieren -> Importeer Excel',
        '',
        'AUTO-DETECTIE VRAAGTYPES:',
        ' - "Meet ..." -> meting (numeriek)',
        ' - "Documenteer ..." / "Noteer ..." -> tekst',
        ' - "Controleer X op werking/aansluiting" -> goed_fout',
        ' - Andere "Inspecteer ..." / "Controleer ..." -> conditiescore (NEN 2767)',
    ]
    for i, line in enumerate(instr, start=2):
        ws_i.cell(row=i, column=1, value=line).font = Font(name='Arial', size=10)
    ws_i.column_dimensions['A'].width = 80

    # Sheet: Templates
    ws_t = wb_out.create_sheet('Templates')
    ws_t.append(['Template naam', 'Asset code (default)', 'Categorie', 'Frequentie', 'Locatie', 'Beschrijving'])
    for cell in ws_t[1]:
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='07567F')
    for tpl_name in templates.keys():
        ws_t.append([tpl_name, '', 'onderhoud', '1 jaarlijks', 'Pompgroepen Den Oever', ''])
    ws_t.column_dimensions['A'].width = 60
    ws_t.column_dimensions['B'].width = 22
    ws_t.column_dimensions['E'].width = 30

    # Sheet: Vragen
    ws_q = wb_out.create_sheet('Vragen')
    headers_q = ['Template naam', 'Sectie volgorde', 'Sectie titel', 'Vraag volgorde', 'Vraagtekst', 'Type', 'Component', 'Discipline', 'Eenheid', 'Permit vereist', 'Verplicht']
    ws_q.append(headers_q)
    for cell in ws_q[1]:
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='07567F')

    for tpl_name, secs in templates.items():
        for si, (sec_title, qs) in enumerate(secs.items()):
            for qi, r in enumerate(qs):
                taak = normalize(r.get('Taakomschrijving') or '')
                qtype = detect_question_type(taak)
                unit = detect_unit(taak) if qtype == 'meting' else ''
                component = normalize(r.get('Bouwdeel specifiek') or r.get('Bouwdeel generiek') or '')
                discipline = (r.get('Discipline') or '').strip()
                if discipline.lower() == 'alle':
                    discipline = ''
                ws_q.append([
                    tpl_name,
                    (si + 1) * 10,
                    sec_title,
                    (qi + 1) * 10,
                    taak,
                    qtype,
                    component,
                    discipline,
                    unit,
                    'nee',
                    'ja',
                ])

    widths = [40, 8, 30, 8, 70, 14, 22, 12, 10, 10, 9]
    for i, w in enumerate(widths):
        ws_q.column_dimensions[get_column_letter(i + 1)].width = w

    # Data validation op type kolom (F)
    dv = DataValidation(type='list', formula1='"goed_fout,conditiescore,meting,tekst"', allow_blank=False)
    ws_q.add_data_validation(dv)
    dv.add(f'F2:F{ws_q.max_row + 100}')

    ws_q.freeze_panes = 'A2'

    wb_out.save(output_path)
    print(f"\nKlaar: {output_path}")
    print(f"  {len(templates)} templates, {sum(sum(len(qs) for qs in secs.values()) for secs in templates.values())} vragen")


if __name__ == '__main__':
    main()
