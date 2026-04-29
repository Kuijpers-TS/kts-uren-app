"""
Genereer Supabase SQL UPDATE/INSERT statements uit de Excel-template.
Sectie-volgordes worden hernummerd naar 10, 20, 30 (in volgorde van de Excel)
zodat tegels in de app correct gesorteerd zijn.

Output: docs/import_inspecties.sql
"""

import json
import openpyxl
from collections import defaultdict


def main():
    wb = openpyxl.load_workbook('docs/KTS_Inspectie_Templates.xlsx', data_only=True)

    # === Templates metadata ===
    ws_t = wb['Templates']
    rows = list(ws_t.iter_rows(values_only=True))
    headers = rows[0]
    templates_meta = {}
    for r in rows[1:]:
        if not r[0]:
            continue
        name = str(r[0]).strip()
        templates_meta[name] = {
            'asset': str(r[1]).strip() if r[1] else None,
            'category': str(r[2]).strip() if r[2] else None,
            'frequency': str(r[3]).strip() if r[3] else None,
            'location': str(r[4]).strip() if r[4] else None,
            'description': str(r[5]).strip() if r[5] else None,
        }

    # === Vragen ===
    ws_q = wb['Vragen']
    qrows = list(ws_q.iter_rows(values_only=True))
    qheaders = qrows[0]
    # Verwacht: Template naam, Sectie volgorde, Sectie titel, Vraag volgorde, Vraagtekst, Type, Component, Discipline, Eenheid, Permit, Verplicht
    # Behoud Excel-volgorde — dat is de natuurlijke volgorde
    sections_per_template = defaultdict(list)  # tpl -> list of {key:(orig_order, title), questions:[]}
    section_index = {}  # (tpl, orig_order, title) -> idx in list

    for r in qrows[1:]:
        if not r[0] or not r[4]:
            continue
        tpl = str(r[0]).strip()
        sec_order = r[1] if r[1] is not None else 0
        sec_title = str(r[2]).strip() if r[2] else ''
        q_order = r[3] if r[3] is not None else 0
        q_text = str(r[4]).strip()
        q_type = str(r[5]).strip() if r[5] else 'goed_fout'
        component = str(r[6]).strip() if r[6] else ''
        discipline = str(r[7]).strip() if r[7] else ''
        unit = str(r[8]).strip() if r[8] else ''
        permit = str(r[9] or 'nee').strip().lower()
        required = str(r[10] or 'ja').strip().lower()

        key = (sec_order, sec_title)
        if (tpl, sec_order, sec_title) not in section_index:
            section_index[(tpl, sec_order, sec_title)] = len(sections_per_template[tpl])
            sections_per_template[tpl].append({
                'orig_order': sec_order,
                'title': sec_title,
                'questions': []
            })
        sec_idx = section_index[(tpl, sec_order, sec_title)]

        question = {
            'text': q_text,
            'type': q_type,
        }
        if component: question['component'] = component
        if discipline: question['discipline'] = discipline
        if unit: question['unit'] = unit
        question['permit_required'] = (permit == 'ja')
        if required == 'nee':
            question['required'] = False

        sections_per_template[tpl][sec_idx]['questions'].append((q_order, question))

    # Sorteer secties per template op orig_order, daarna op insertvolgorde (stable)
    # Vragen sorteren binnen sectie op q_order
    output_templates = {}
    for tpl, secs in sections_per_template.items():
        # Stable sort op orig_order zodat dubbele orders Excel-volgorde behouden
        secs_sorted = sorted(enumerate(secs), key=lambda x: (x[1]['orig_order'], x[0]))
        new_sections = []
        for new_idx, (_, sec) in enumerate(secs_sorted):
            sec['questions'].sort(key=lambda x: x[0])
            new_sections.append({
                'id': f's{(new_idx + 1) * 10}',
                'title': sec['title'],
                'questions': [q for _, q in sec['questions']]
            })
        output_templates[tpl] = new_sections

    # === Genereer SQL ===
    sql_lines = []
    sql_lines.append("-- =================================================================")
    sql_lines.append("-- KTS Inspectie templates import — gegenereerd uit Excel")
    sql_lines.append("-- =================================================================")
    sql_lines.append("-- Voer dit uit in Supabase → SQL Editor")
    sql_lines.append("-- Bestaande templates worden ge-UPSERT (vervangen op naam)")
    sql_lines.append("-- =================================================================")
    sql_lines.append("")

    def sql_str_or_null(v):
        if not v:
            return 'NULL'
        return "'" + str(v).replace("'", "''") + "'"

    for tpl_name, sections in output_templates.items():
        meta = templates_meta.get(tpl_name, {})
        sections_json = json.dumps(sections, ensure_ascii=False)
        sections_sql = sections_json.replace("'", "''")
        name_sql = tpl_name.replace("'", "''")
        asset_v = sql_str_or_null(meta.get('asset'))
        cat_v = sql_str_or_null(meta.get('category'))
        freq_v = sql_str_or_null(meta.get('frequency'))
        loc_v = sql_str_or_null(meta.get('location'))
        desc_v = sql_str_or_null(meta.get('description'))

        total_q = sum(len(s['questions']) for s in sections)
        sql_lines.append(f"-- Template: {tpl_name} ({len(sections)} secties, {total_q} vragen)")
        sql_lines.append("DO $$")
        sql_lines.append("BEGIN")
        sql_lines.append(f"    IF EXISTS (SELECT 1 FROM inspection_templates WHERE name = '{name_sql}') THEN")
        sql_lines.append("        UPDATE inspection_templates SET")
        sql_lines.append(f"            asset = {asset_v},")
        sql_lines.append(f"            category = {cat_v},")
        sql_lines.append(f"            frequency = {freq_v},")
        sql_lines.append(f"            location = {loc_v},")
        sql_lines.append(f"            description = {desc_v},")
        sql_lines.append(f"            sections = '{sections_sql}'::jsonb,")
        sql_lines.append("            is_active = true")
        sql_lines.append(f"        WHERE name = '{name_sql}';")
        sql_lines.append("    ELSE")
        sql_lines.append("        INSERT INTO inspection_templates (name, asset, category, frequency, location, description, sections, is_active)")
        sql_lines.append("        VALUES (")
        sql_lines.append(f"            '{name_sql}',")
        sql_lines.append(f"            {asset_v},")
        sql_lines.append(f"            {cat_v},")
        sql_lines.append(f"            {freq_v},")
        sql_lines.append(f"            {loc_v},")
        sql_lines.append(f"            {desc_v},")
        sql_lines.append(f"            '{sections_sql}'::jsonb,")
        sql_lines.append("            true")
        sql_lines.append("        );")
        sql_lines.append("    END IF;")
        sql_lines.append("END $$;")
        sql_lines.append("")

    output = '\n'.join(sql_lines)
    with open('docs/import_inspecties.sql', 'w', encoding='utf-8') as f:
        f.write(output)

    print("SQL gegenereerd: docs/import_inspecties.sql")
    print()
    for tpl_name, sections in output_templates.items():
        total_q = sum(len(s['questions']) for s in sections)
        print(f"  {tpl_name}")
        print(f"    {len(sections)} secties, {total_q} vragen")
        for s in sections:
            print(f"      [{s['id']}] {s['title']} ({len(s['questions'])} vragen)")
        print()


if __name__ == '__main__':
    main()
