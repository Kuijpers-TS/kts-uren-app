"""
Genereer een Excel template voor KTS inspecties.
Bevat de twee bestaande templates (beunkoelers/ruwwaterpompen + koelwater skids)
plus een instructiesheet en data validation dropdowns.

Run: python docs/generate_inspectie_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# === DATA VAN BESTAANDE TEMPLATES ===
templates = {
    "Visuele inspectie beunkoelers & ruwwaterpompen": {
        "asset_default": "NSM-PG1",
        "category": "onderhoud",
        "frequency": "1 jaarlijks",
        "location": "Pompgroepen Den Oever",
        "sections": [
            {"title": "PG1 — Ruwwaterpomp 1", "questions": [
                ("Ruwwaterpomp op aangroei en beschadigingen controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Spervloeistof controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Elektromotor op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Motordeksel op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Kabeldoorvoer en kabel op beschadigingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Weerstandsvrijheid van aansluitingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Kabels, doorvoeren, IP68 afdichtingen en hijsogen controleren", "conditiescore", "Bekabeling", "WTB"),
            ]},
            {"title": "PG1 — Ruwwaterpomp 2", "questions": [
                ("Ruwwaterpomp op aangroei en beschadigingen controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Spervloeistof controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Elektromotor op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Motordeksel op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Kabeldoorvoer en kabel op beschadigingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Weerstandsvrijheid van aansluitingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Kabels, doorvoeren, IP68 afdichtingen en hijsogen controleren", "conditiescore", "Bekabeling", "WTB"),
            ]},
            {"title": "PG1 — Beunkoeler 1", "questions": [
                ("Beunkoeler op aangroei en beschadigingen controleren", "conditiescore", "Beunkoeler", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Load cells indicator reinigen met droge doek", "conditiescore", "Sensor", "WTB"),
            ]},
            {"title": "PG1 — Beunkoeler 2", "questions": [
                ("Beunkoeler op aangroei en beschadigingen controleren", "conditiescore", "Beunkoeler", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Load cells indicator reinigen met droge doek", "conditiescore", "Sensor", "WTB"),
            ]},
            {"title": "PG2 — Ruwwaterpomp 1", "questions": [
                ("Ruwwaterpomp op aangroei en beschadigingen controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Spervloeistof controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Elektromotor op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Motordeksel op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Kabeldoorvoer en kabel op beschadigingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Weerstandsvrijheid van aansluitingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Kabels, doorvoeren, IP68 afdichtingen en hijsogen controleren", "conditiescore", "Bekabeling", "WTB"),
            ]},
            {"title": "PG2 — Ruwwaterpomp 2", "questions": [
                ("Ruwwaterpomp op aangroei en beschadigingen controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Spervloeistof controleren", "conditiescore", "Ruwwaterpomp", "WTB"),
                ("Elektromotor op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Motordeksel op beschadigingen controleren", "conditiescore", "Elektromotor", "WTB"),
                ("Kabeldoorvoer en kabel op beschadigingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Weerstandsvrijheid van aansluitingen controleren", "conditiescore", "Bekabeling", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Kabels, doorvoeren, IP68 afdichtingen en hijsogen controleren", "conditiescore", "Bekabeling", "WTB"),
            ]},
            {"title": "PG2 — Beunkoeler 1", "questions": [
                ("Beunkoeler op aangroei en beschadigingen controleren", "conditiescore", "Beunkoeler", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Load cells indicator reinigen met droge doek", "conditiescore", "Sensor", "WTB"),
            ]},
            {"title": "PG2 — Beunkoeler 2", "questions": [
                ("Beunkoeler op aangroei en beschadigingen controleren", "conditiescore", "Beunkoeler", "WTB"),
                ("Leidingen, flenzen en appendages op lekkage en corrosie controleren", "conditiescore", "Buisleiding", "WTB"),
                ("Load cells indicator reinigen met droge doek", "conditiescore", "Sensor", "WTB"),
            ]},
            {"title": "Algemeen — Inlaat & sensoren", "questions": [
                ("Inlaatroosters inspecteren en grofvuil verwijderen", "conditiescore", "Inlaatrooster", "WTB"),
                ("Zuigkast inspecteren en grofvuil verwijderen", "conditiescore", "Zuigkast", "WTB"),
                ("Sensoren schoonmaken", "conditiescore", "Sensor", "WTB"),
            ]},
        ]
    },
    "Visuele inspectie koelwater skids": {
        "asset_default": "",
        "category": "onderhoud",
        "frequency": "1 jaarlijks",
        "location": "",
        "sections": [
            {"title": "Hoofdpomp skid", "questions": [
                ("Controleer op lekkages/beschadigingen aan leidingen", "goed_fout", "Buisleiding", "WTB"),
                ("Controleer pompen op lekkages en beschadigingen", "goed_fout", "Pomp", "WTB"),
                ("Controleer warmtewisselaars op lekkages en beschadigingen", "goed_fout", "Warmtewisselaar", "WTB"),
                ("Controleer op corrosie/vervuiling van metalen delen (pompen, frames, afsluiters en flenzen)", "goed_fout", "Frame", "WTB"),
                ("Controleer kabels en aansluitingen op losse verbindingen of slijtage", "goed_fout", "Bekabeling", "IA&E"),
                ("Controleer debietmeters op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer temperatuuropnemers op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer overige sensoren op losse verbindingen of slijtage", "goed_fout", "Sensor", "IA&E"),
                ("Controleer motoren op beschadiging", "goed_fout", "Elektromotor", "WTB"),
                ("Controleer motoren op trillingen, warmteontwikkeling of afwijkingen", "goed_fout", "Elektromotor", "WTB"),
            ]},
            {"title": "Verdelerskid 1 (Lagers)", "questions": [
                ("Controleer op lekkages/beschadigingen aan leidingen", "goed_fout", "Buisleiding", "WTB"),
                ("Controleer warmtewisselaars op lekkages en beschadigingen", "goed_fout", "Warmtewisselaar", "WTB"),
                ("Controleer op corrosie/vervuiling van metalen delen (pompen, frames, afsluiters en flenzen)", "goed_fout", "Frame", "WTB"),
                ("Controleer kabels en aansluitingen op losse verbindingen of slijtage", "goed_fout", "Bekabeling", "IA&E"),
                ("Controleer debietmeters op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer temperatuuropnemers op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer motoren op beschadiging, trillingen en warmteontwikkeling", "goed_fout", "Elektromotor", "WTB"),
            ]},
            {"title": "Verdelerskid 2 (Motor)", "questions": [
                ("Controleer op lekkages/beschadigingen aan leidingen", "goed_fout", "Buisleiding", "WTB"),
                ("Controleer warmtewisselaars op lekkages en beschadigingen", "goed_fout", "Warmtewisselaar", "WTB"),
                ("Controleer op corrosie/vervuiling van metalen delen (pompen, frames, afsluiters en flenzen)", "goed_fout", "Frame", "WTB"),
                ("Controleer kabels en aansluitingen op losse verbindingen of slijtage", "goed_fout", "Bekabeling", "IA&E"),
                ("Controleer debietmeters op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer temperatuuropnemers op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer motoren op beschadiging, trillingen en warmteontwikkeling", "goed_fout", "Elektromotor", "WTB"),
            ]},
            {"title": "Verdelerskid 3 (FO)", "questions": [
                ("Controleer op lekkages/beschadigingen aan leidingen", "goed_fout", "Buisleiding", "WTB"),
                ("Controleer warmtewisselaars op lekkages en beschadigingen", "goed_fout", "Warmtewisselaar", "WTB"),
                ("Controleer op corrosie/vervuiling van metalen delen (pompen, frames, afsluiters en flenzen)", "goed_fout", "Frame", "WTB"),
                ("Controleer kabels en aansluitingen op losse verbindingen of slijtage", "goed_fout", "Bekabeling", "IA&E"),
                ("Controleer debietmeters op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer temperatuuropnemers op werking en aansluiting", "goed_fout", "Sensor", "IA&E"),
                ("Controleer motoren op beschadiging, trillingen en warmteontwikkeling", "goed_fout", "Elektromotor", "WTB"),
            ]},
        ]
    }
}


# === STIJL CONSTANTEN ===
KTS_BLUE = "07567F"
KTS_BLUE_LIGHT = "E0F2F7"
HEADER_BG = "07567F"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "F8FAFC"
BORDER_COLOR = "E2E8F0"

thin = Side(border_style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name="Arial", size=11, bold=True, color=HEADER_FG)
body_font = Font(name="Arial", size=10)
mono_font = Font(name="Consolas", size=10)
section_font = Font(name="Arial", size=11, bold=True, color=KTS_BLUE)
title_font = Font(name="Arial", size=14, bold=True, color=KTS_BLUE)


def create_workbook():
    wb = Workbook()

    # Sheet 1: Instructies
    ws_inst = wb.active
    ws_inst.title = "Instructies"
    create_instructions(ws_inst)

    # Sheet 2: Templates (metadata)
    ws_tpl = wb.create_sheet("Templates")
    create_templates_sheet(ws_tpl)

    # Sheet 3: Vragen (alle vragen plat)
    ws_q = wb.create_sheet("Vragen")
    create_questions_sheet(ws_q)

    # Sheet 4: Lookups (verborgen, voor data validation)
    ws_look = wb.create_sheet("Lookups")
    create_lookups_sheet(ws_look)
    ws_look.sheet_state = "hidden"

    return wb


def create_instructions(ws):
    ws["A1"] = "KTS Inspectie Templates — Excel import"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    intro = [
        "",
        "Deze Excel bevat de inspectie-templates die in de KTS Uren App staan.",
        "Je kunt deze aanpassen en daarna terug importeren in de app.",
        "",
        "STRUCTUUR:",
        "  • Sheet 'Templates' — metadata per template (naam, asset, frequentie, locatie, etc.)",
        "  • Sheet 'Vragen' — één rij per vraag, gegroepeerd per template + sectie",
        "  • Sheet 'Lookups' — verborgen, bevat dropdownopties (niet aanpassen)",
        "",
        "WERKWIJZE — VRAGEN AANPASSEN:",
        "  1. Open sheet 'Vragen'",
        "  2. Pas 'Vraagtekst' aan, of voeg/verwijder rijen",
        "  3. Hou kolom 'Template naam' en 'Sectie titel' consistent met sheet 'Templates'",
        "  4. Sectie volgorde en Vraag volgorde bepalen de plaats — gebruik 10, 20, 30 ipv 1, 2, 3 zodat je makkelijk tussen kunt voegen",
        "",
        "VRAAGTYPES:",
        "  • goed_fout — Goed / Fout / N.v.t. knoppen",
        "  • conditiescore — NEN 2767 schaal 1 (uitstekend) t/m 6 (zeer slecht)",
        "  • meting — numeriek invoerveld met eenheid",
        "  • tekst — vrij tekstveld",
        "",
        "DISCIPLINES (afkortingen):",
        "  • WTB — Werktuigbouw",
        "  • IA&E — Industriële Automatisering & Elektrotechniek",
        "  • PB — Procesbesturing",
        "",
        "NIEUWE TEMPLATES TOEVOEGEN:",
        "  1. Voeg rij toe in sheet 'Templates' met naam + metadata",
        "  2. Voeg vragen toe in sheet 'Vragen' met die exact dezelfde template-naam",
        "",
        "IMPORT IN DE APP:",
        "  Beheer → Formulieren → 📥 Importeren uit Excel (volgt nog in v2)",
        "",
        "EXPORT VANUIT DE APP:",
        "  Beheer → Formulieren → 📤 Exporteren naar Excel (volgt nog in v2)",
        "",
        "VRAGEN OF PROBLEMEN: vraag het Mark.",
    ]
    for i, line in enumerate(intro, start=2):
        cell = ws[f"A{i}"]
        cell.value = line
        if line.endswith(":"):
            cell.font = Font(name="Arial", size=10, bold=True, color=KTS_BLUE)
        else:
            cell.font = body_font
    ws.column_dimensions["A"].width = 100


def create_templates_sheet(ws):
    headers = ["Template naam", "Asset code (default)", "Categorie", "Frequentie", "Locatie", "Beschrijving"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row = 2
    for name, meta in templates.items():
        ws.cell(row=row, column=1, value=name).font = body_font
        ws.cell(row=row, column=2, value=meta["asset_default"]).font = mono_font
        ws.cell(row=row, column=3, value=meta["category"]).font = body_font
        ws.cell(row=row, column=4, value=meta["frequency"]).font = body_font
        ws.cell(row=row, column=5, value=meta["location"]).font = body_font
        ws.cell(row=row, column=6, value="").font = body_font
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    # Kolombreedtes
    widths = [50, 22, 16, 16, 28, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.row_dimensions[1].height = 24

    # Freeze panes
    ws.freeze_panes = "A2"


def create_questions_sheet(ws):
    headers = [
        "Template naam",
        "Sectie volgorde",
        "Sectie titel",
        "Vraag volgorde",
        "Vraagtekst",
        "Type",
        "Component",
        "Discipline",
        "Eenheid",
        "Permit vereist",
        "Verplicht",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row = 2
    for tpl_name, meta in templates.items():
        for sec_idx, sec in enumerate(meta["sections"]):
            sec_order = (sec_idx + 1) * 10  # 10, 20, 30 voor makkelijk tussenvoegen
            for q_idx, q in enumerate(sec["questions"]):
                q_order = (q_idx + 1) * 10
                text, qtype, component, discipline = q
                ws.cell(row=row, column=1, value=tpl_name).font = body_font
                ws.cell(row=row, column=2, value=sec_order).font = mono_font
                ws.cell(row=row, column=3, value=sec["title"]).font = body_font
                ws.cell(row=row, column=4, value=q_order).font = mono_font
                ws.cell(row=row, column=5, value=text).font = body_font
                ws.cell(row=row, column=6, value=qtype).font = mono_font
                ws.cell(row=row, column=7, value=component).font = body_font
                ws.cell(row=row, column=8, value=discipline).font = mono_font
                ws.cell(row=row, column=9, value="").font = mono_font
                ws.cell(row=row, column=10, value="nee").font = mono_font
                ws.cell(row=row, column=11, value="ja").font = mono_font
                # Alterneer rij-achtergrond per template
                bg = ALT_ROW_BG if list(templates.keys()).index(tpl_name) % 2 == 1 else None
                for col in range(1, 12):
                    c = ws.cell(row=row, column=col)
                    c.border = border
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    if bg:
                        c.fill = PatternFill("solid", start_color=bg)
                row += 1

    # Kolombreedtes
    widths = [40, 8, 28, 8, 60, 14, 18, 12, 10, 10, 9]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.row_dimensions[1].height = 24

    # Data validation: Type kolom (F)
    dv_type = DataValidation(type="list", formula1='"goed_fout,conditiescore,meting,tekst"', allow_blank=False)
    dv_type.error = "Kies één van: goed_fout, conditiescore, meting, tekst"
    dv_type.errorTitle = "Ongeldig type"
    ws.add_data_validation(dv_type)
    dv_type.add(f"F2:F{row + 100}")  # ruimte voor toevoegen

    # Discipline kolom (H)
    dv_disc = DataValidation(type="list", formula1='"WTB,IA&E,PB"', allow_blank=True)
    ws.add_data_validation(dv_disc)
    dv_disc.add(f"H2:H{row + 100}")

    # Permit (J) en Verplicht (K)
    dv_yn = DataValidation(type="list", formula1='"ja,nee"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"J2:J{row + 100}")
    dv_yn.add(f"K2:K{row + 100}")

    # Freeze panes (header + eerste 3 kolommen)
    ws.freeze_panes = "D2"


def create_lookups_sheet(ws):
    ws["A1"] = "Vraagtypes"
    ws["B1"] = "Disciplines"
    ws["C1"] = "Ja/Nee"
    types = ["goed_fout", "conditiescore", "meting", "tekst"]
    disc = ["WTB", "IA&E", "PB"]
    yn = ["ja", "nee"]
    for i, t in enumerate(types, start=2):
        ws.cell(row=i, column=1, value=t)
    for i, d in enumerate(disc, start=2):
        ws.cell(row=i, column=2, value=d)
    for i, v in enumerate(yn, start=2):
        ws.cell(row=i, column=3, value=v)


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "docs/KTS_Inspectie_Templates.xlsx"
    wb.save(output_path)
    print(f"Excel template geschreven: {output_path}")

    # Stats
    total_q = sum(sum(len(s["questions"]) for s in t["sections"]) for t in templates.values())
    total_s = sum(len(t["sections"]) for t in templates.values())
    print(f"Aantal templates: {len(templates)}")
    print(f"Aantal secties: {total_s}")
    print(f"Aantal vragen: {total_q}")
