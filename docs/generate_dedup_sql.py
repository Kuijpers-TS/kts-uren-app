"""
Generate SQL die templates ontdubbelt: PG1/PG2 prefix verwijderen,
dubbele secties wegmieteren. Resultaat: 1 template = 1 inspectie
voor 1 asset (NSM-PG1 of NSM-PG2 via asset-code bij start).
"""

import json
import openpyxl
from collections import OrderedDict

PREFIXES_TO_STRIP = ['PG1 — ', 'PG2 — ', 'PG1 - ', 'PG2 - ']


def strip_pg_prefix(title: str) -> str:
    for p in PREFIXES_TO_STRIP:
        if title.startswith(p):
            return title[len(p):]
    return title


def main():
    wb = openpyxl.load_workbook('docs/KTS_Inspectie_Templates.xlsx', data_only=True)

    # Templates metadata
    ws_t = wb['Templates']
    rows_t = list(ws_t.iter_rows(values_only=True))
    templates_meta = {}
    for r in rows_t[1:]:
        if not r[0]:
            continue
        templates_meta[str(r[0]).strip()] = {
            'asset': str(r[1]).strip() if r[1] else None,
            'category': str(r[2]).strip() if r[2] else None,
            'frequency': str(r[3]).strip() if r[3] else None,
            'location': str(r[4]).strip() if r[4] else None,
            'description': str(r[5]).strip() if r[5] else None,
        }

    # Vragen
    ws_q = wb['Vragen']
    rows_q = list(ws_q.iter_rows(values_only=True))

    # Groepeer per (template, gestripte sectie-titel) -> behoud Excel-volgorde
    # Skip duplicaten: zelfde (tpl, gestripte_titel, vraagtekst) komt 1x voor
    templates_data = OrderedDict()
    seen_questions = set()

    for r in rows_q[1:]:
        if not r[0] or not r[4]:
            continue
        tpl = str(r[0]).strip()
        sec_title_raw = str(r[2] or '').strip()
        sec_title = strip_pg_prefix(sec_title_raw)
        q_text = str(r[4]).strip()
        q_type = str(r[5] or 'goed_fout').strip()
        component = str(r[6] or '').strip()
        discipline = str(r[7] or '').strip()
        unit = str(r[8] or '').strip()
        permit = str(r[9] or 'nee').strip().lower()
        required = str(r[10] or 'ja').strip().lower()

        if tpl not in templates_data:
            templates_data[tpl] = OrderedDict()
        if sec_title not in templates_data[tpl]:
            templates_data[tpl][sec_title] = []

        # Dedupe: skip als zelfde vraag al in deze sectie zit
        dedupe_key = (tpl, sec_title, q_text, component)
        if dedupe_key in seen_questions:
            continue
        seen_questions.add(dedupe_key)

        question = {'text': q_text, 'type': q_type}
        if component: question['component'] = component
        if discipline: question['discipline'] = discipline
        if unit: question['unit'] = unit
        question['permit_required'] = (permit == 'ja')
        if required == 'nee':
            question['required'] = False

        templates_data[tpl][sec_title].append(question)

    # Bouw section arrays met volgorde 10/20/30
    output = OrderedDict()
    for tpl, secs in templates_data.items():
        sections = []
        for i, (title, qs) in enumerate(secs.items()):
            sections.append({
                'id': f's{(i + 1) * 10}',
                'title': title,
                'questions': qs
            })
        output[tpl] = sections

    # SQL genereren met DO-blocks
    def sql_str_or_null(v):
        if not v:
            return 'NULL'
        return "'" + str(v).replace("'", "''") + "'"

    sql_lines = []
    sql_lines.append("-- =================================================================")
    sql_lines.append("-- Dedup templates: PG1/PG2 prefix uit secties weg, dubbele vragen weg")
    sql_lines.append("-- 1 template = 1 inspectie voor 1 asset (NSM-PG1 of NSM-PG2 via asset-code)")
    sql_lines.append("-- =================================================================")
    sql_lines.append("")

    for tpl_name, sections in output.items():
        meta = templates_meta.get(tpl_name, {})
        sections_json = json.dumps(sections, ensure_ascii=False)
        sections_sql = sections_json.replace("'", "''")
        name_sql = tpl_name.replace("'", "''")
        total_q = sum(len(s['questions']) for s in sections)
        sql_lines.append(f"-- {tpl_name} ({len(sections)} secties, {total_q} vragen)")
        sql_lines.append("DO $$")
        sql_lines.append("BEGIN")
        sql_lines.append(f"    IF EXISTS (SELECT 1 FROM inspection_templates WHERE name = '{name_sql}') THEN")
        sql_lines.append("        UPDATE inspection_templates SET")
        sql_lines.append(f"            asset = {sql_str_or_null(meta.get('asset'))},")
        sql_lines.append(f"            category = {sql_str_or_null(meta.get('category'))},")
        sql_lines.append(f"            frequency = {sql_str_or_null(meta.get('frequency'))},")
        sql_lines.append(f"            location = {sql_str_or_null(meta.get('location'))},")
        sql_lines.append(f"            description = {sql_str_or_null(meta.get('description'))},")
        sql_lines.append(f"            sections = '{sections_sql}'::jsonb,")
        sql_lines.append("            is_active = true")
        sql_lines.append(f"        WHERE name = '{name_sql}';")
        sql_lines.append("    END IF;")
        sql_lines.append("END $$;")
        sql_lines.append("")

    with open('docs/dedup_inspecties.sql', 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print("SQL: docs/dedup_inspecties.sql")
    print()
    for tpl, sections in output.items():
        total_q = sum(len(s['questions']) for s in sections)
        print(f"  {tpl}: {len(sections)} secties, {total_q} vragen")
        for s in sections:
            print(f"    [{s['id']}] {s['title']} ({len(s['questions'])} vragen)")
        print()


if __name__ == '__main__':
    main()
